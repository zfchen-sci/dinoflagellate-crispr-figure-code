from __future__ import annotations

import argparse
import json
from itertools import combinations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from scipy import stats
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from statsmodels.stats.multitest import multipletests
from statsmodels.stats.oneway import anova_oneway

from common import (
    format_p,
    games_howell,
    infer_fig4_qpcr_strain,
    mean_sd,
    significance_label,
)


WT = "Wild-type strain"
EDITED = "sxtA4-edited strain"


def t_row(
    figure: str,
    panel: str,
    comparison: str,
    outcome: str,
    a,
    b,
    test_name: str,
    equal_var: bool,
    source_file: str,
    source_sheet: str,
    correction: str = "None",
    p_adjusted: float | None = None,
    tiers: int = 4,
    notes: str = "",
) -> dict:
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    a = a[np.isfinite(a)]
    b = b[np.isfinite(b)]
    result = stats.ttest_ind(a, b, equal_var=equal_var)
    p_display = float(result.pvalue if p_adjusted is None else p_adjusted)
    ma, sda, na = mean_sd(a)
    mb, sdb, nb = mean_sd(b)
    if equal_var:
        df = na + nb - 2
    else:
        va = np.var(a, ddof=1) / na
        vb = np.var(b, ddof=1) / nb
        df = (va + vb) ** 2 / (va**2 / (na - 1) + vb**2 / (nb - 1))
    return {
        "figure": figure,
        "panel": panel,
        "comparison": comparison,
        "outcome": outcome,
        "group_1_mean": ma,
        "group_1_sd": sda,
        "group_1_n": na,
        "group_2_mean": mb,
        "group_2_sd": sdb,
        "group_2_n": nb,
        "test": test_name,
        "correction": correction,
        "statistic": float(result.statistic),
        "df": float(df),
        "p_raw": float(result.pvalue),
        "p_adjusted": p_adjusted,
        "p_used_for_display": p_display,
        "p_formatted": format_p(p_display),
        "significance": significance_label(p_display, tiers=tiers),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "notes": notes,
    }


def add_holm(rows: list[dict], indices: list[int]) -> None:
    adjusted = multipletests([rows[i]["p_raw"] for i in indices], method="holm")[1]
    for i, p_adj in zip(indices, adjusted):
        rows[i]["p_adjusted"] = float(p_adj)
        rows[i]["p_used_for_display"] = float(p_adj)
        rows[i]["p_formatted"] = format_p(float(p_adj))
        rows[i]["significance"] = significance_label(float(p_adj), tiers=4)
        rows[i]["correction"] = "Holm"


def parse_injection_data(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Fig.3b-g_raw", header=21)
    df = df[df["crRNA"].isin(["Site1", "Site2"])].copy()
    for label, denominator in {
        "per_injected": "Injected cells",
        "per_germinated": "Germinated cells",
        "per_viable": "Viable cells",
    }.items():
        df[label] = 100 * df["Positive cells"] / df[denominator].replace(0, np.nan)
    return df


def calculate_statistics(source_dir: Path) -> dict[str, list[dict]]:
    tests: list[dict] = []
    summaries: list[dict] = []
    checks: list[dict] = []

    f2 = source_dir / "Source_data_fig2.xlsx"
    d2 = pd.read_excel(f2, sheet_name="Fig.2d")
    for outcome in ("germination_rate", "viability_rate"):
        tests.append(
            t_row(
                "Fig. 2",
                "d",
                "500-nm tip vs 50-nm tip",
                outcome,
                d2.loc[d2.tip_size_nm == 500, outcome],
                d2.loc[d2.tip_size_nm == 50, outcome],
                "Two-tailed Welch t-test",
                False,
                f2.name,
                "Fig.2d",
                tiers=1,
            )
        )
    d2f = pd.read_excel(f2, sheet_name="Fig.2f")
    for outcome in ("germination_rate", "viability_rate"):
        tests.append(
            t_row(
                "Fig. 2",
                "f",
                "0.003 µM vs 0.012 µM UvrD",
                outcome,
                d2f.loc[np.isclose(d2f.uvrd_uM, 0.003), outcome],
                d2f.loc[np.isclose(d2f.uvrd_uM, 0.012), outcome],
                "Two-tailed Welch t-test",
                False,
                f2.name,
                "Fig.2f",
                tiers=1,
            )
        )

    f3 = source_dir / "Source_data_fig3.xlsx"
    inj = parse_injection_data(f3)
    b = inj[(inj.crRNA == "Site2") & (inj["Injection strategy"] == "UvrD+Nuclei")]
    groups = [b.loc[b["RNP concentration (nM)"] == dose, "per_viable"].dropna() for dose in (5, 30, 1200)]
    omnibus = stats.f_oneway(*groups)
    tests.append(
        {
            "figure": "Fig. 3",
            "panel": "f",
            "comparison": "5 vs 30 vs 1,200 nM RNP",
            "outcome": "editing recovery per viable cell",
            "group_1_mean": None,
            "group_1_sd": None,
            "group_1_n": 3,
            "group_2_mean": None,
            "group_2_sd": None,
            "group_2_n": 3,
            "test": "One-way ANOVA",
            "correction": "Tukey HSD for pairwise comparisons",
            "statistic": float(omnibus.statistic),
            "df": "2, 6",
            "p_raw": float(omnibus.pvalue),
            "p_adjusted": None,
            "p_used_for_display": float(omnibus.pvalue),
            "p_formatted": format_p(float(omnibus.pvalue)),
            "significance": significance_label(float(omnibus.pvalue)),
            "source_file": f3.name,
            "source_sheet": "Fig.3b-g_raw",
            "notes": "Omnibus test; the figure displays Tukey pairwise results.",
        }
    )
    tukey_b = pairwise_tukeyhsd(b["per_viable"], b["RNP concentration (nM)"])
    for (g1, g2), diff, p_adj, ci in zip(
        combinations(tukey_b.groupsunique, 2), tukey_b.meandiffs, tukey_b.pvalues, tukey_b.confint
    ):
        tests.append(
            {
                "figure": "Fig. 3",
                "panel": "f",
                "comparison": f"{g1:g} vs {g2:g} nM RNP",
                "outcome": "editing recovery per viable cell",
                "group_1_mean": float(b.loc[b["RNP concentration (nM)"] == g1, "per_viable"].mean()),
                "group_1_sd": float(b.loc[b["RNP concentration (nM)"] == g1, "per_viable"].std(ddof=1)),
                "group_1_n": 3,
                "group_2_mean": float(b.loc[b["RNP concentration (nM)"] == g2, "per_viable"].mean()),
                "group_2_sd": float(b.loc[b["RNP concentration (nM)"] == g2, "per_viable"].std(ddof=1)),
                "group_2_n": 3,
                "test": "Tukey HSD",
                "correction": "Family-wise error rate",
                "statistic": float(diff),
                "df": None,
                "p_raw": None,
                "p_adjusted": float(p_adj),
                "p_used_for_display": float(p_adj),
                "p_formatted": format_p(float(p_adj)),
                "significance": significance_label(float(p_adj)),
                "source_file": f3.name,
                "source_sheet": "Fig.3b-g_raw",
                "notes": f"95% CI for mean difference: {ci[0]:.4f} to {ci[1]:.4f}.",
            }
        )

    c = inj[(inj["Injection strategy"] == "UvrD+Nuclei") & (inj["RNP concentration (nM)"] == 1200)]
    for outcome in ("per_injected", "per_germinated", "per_viable"):
        tests.append(
            t_row(
                "Fig. 3",
                "g",
                "crRNA-site1 vs crRNA-site2",
                outcome,
                c.loc[c.crRNA == "Site1", outcome],
                c.loc[c.crRNA == "Site2", outcome],
                "Two-tailed Welch t-test",
                False,
                f3.name,
                "Fig.3b-g_raw",
            )
        )

    edit = pd.read_excel(f3, sheet_name="Fig.3h-j")
    edit["lineage"] = edit["sample"].str.rsplit("-", n=1).str[0]
    metric = "corrected_editing_efficiency_pct"
    lineage = edit.groupby(["Target_site", "lineage"], as_index=False)[metric].mean()
    tests.append(
        t_row(
            "Fig. 3",
            "h",
            "crRNA-site1 vs crRNA-site2",
            "mean corrected editing efficiency per lineage",
            lineage.loc[lineage.Target_site == "crRNA_site1", metric],
            lineage.loc[lineage.Target_site == "crRNA_site2", metric],
            "Two-tailed Welch t-test",
            False,
            f3.name,
            "Fig.3h-j",
            notes="Lineage means are the independent observations; each mean summarizes three sequencing measurements.",
        )
    )
    for site, sub in edit.groupby("Target_site"):
        grouped = {
            str(lineage_name): group[metric].to_numpy()
            for lineage_name, group in sub.groupby("lineage", sort=False)
        }
        omnibus = anova_oneway(tuple(grouped.values()), use_var="unequal", welch_correction=True)
        tests.append(
            {
                "figure": "Fig. 3",
                "panel": "i",
                "comparison": f"All lineages within {site.replace('_', '-')}",
                "outcome": "corrected editing efficiency",
                "group_1_mean": None,
                "group_1_sd": None,
                "group_1_n": 3,
                "group_2_mean": None,
                "group_2_sd": None,
                "group_2_n": 3,
                "test": "Welch one-way ANOVA",
                "correction": "Games-Howell for pairwise comparisons",
                "statistic": float(omnibus.statistic),
                "df": f"{omnibus.df_num:.0f}, {omnibus.df_denom:.6f}",
                "p_raw": float(omnibus.pvalue),
                "p_adjusted": None,
                "p_used_for_display": float(omnibus.pvalue),
                "p_formatted": format_p(float(omnibus.pvalue)),
                "significance": significance_label(float(omnibus.pvalue)),
                "source_file": f3.name,
                "source_sheet": "Fig.3h-j",
                "notes": "Independent biological replicates; variances are not assumed to be equal.",
            }
        )
        for comparison in games_howell(grouped):
            g1, g2 = comparison["group_1"], comparison["group_2"]
            p_adj = comparison["p_adjusted"]
            tests.append(
                {
                    "figure": "Fig. 3",
                    "panel": "i",
                    "comparison": f"{g1} vs {g2}",
                    "outcome": "corrected editing efficiency",
                    "group_1_mean": float(sub.loc[sub.lineage == g1, metric].mean()),
                    "group_1_sd": float(sub.loc[sub.lineage == g1, metric].std(ddof=1)),
                    "group_1_n": 3,
                    "group_2_mean": float(sub.loc[sub.lineage == g2, metric].mean()),
                    "group_2_sd": float(sub.loc[sub.lineage == g2, metric].std(ddof=1)),
                    "group_2_n": 3,
                    "test": "Games-Howell pairwise test",
                    "correction": "Studentized-range family-wise error rate",
                    "statistic": float(comparison["mean_difference"]),
                    "df": float(comparison["df"]),
                    "p_raw": None,
                    "p_adjusted": float(p_adj),
                    "p_used_for_display": float(p_adj),
                    "p_formatted": format_p(float(p_adj)),
                    "significance": significance_label(float(p_adj)),
                    "source_file": f3.name,
                    "source_sheet": "Fig.3h-j",
                    "notes": (
                        "Independent biological replicates; 95% CI for group 2 minus group 1: "
                        f"{comparison['ci_low']:.4f} to {comparison['ci_high']:.4f}."
                    ),
                }
            )
    f4 = source_dir / "Source_data_fig4.xlsx"
    q = pd.read_excel(f4, sheet_name="Fig.4d", header=2)
    qbio = q.groupby(["Samples", "Days", "Cell density (cells/ml)"], as_index=False).Ct.mean()
    qbio["Strain"] = infer_fig4_qpcr_strain(qbio["Samples"], WT, EDITED)
    qbio["transcripts_per_cell"] = 10 ** ((39.67 - qbio.Ct) / 3.59) / qbio["Cell density (cells/ml)"]
    qbio["log10_transcripts_per_cell"] = np.log10(qbio["transcripts_per_cell"])
    fig4d_indices = []
    for day, sub in qbio.groupby("Days"):
        fig4d_indices.append(len(tests))
        tests.append(
            t_row(
                "Fig. 4",
                "d",
                f"WT vs edited at day {day}",
                "log10(sxtA4 transcripts per cell)",
                sub.loc[sub.Strain == WT, "log10_transcripts_per_cell"],
                sub.loc[sub.Strain == EDITED, "log10_transcripts_per_cell"],
                "Two-tailed Welch t-test",
                False,
                f4.name,
                "Fig.4d",
                notes="Technical Ct replicates were averaged within each biological sample before conversion; the test uses log10 abundance and the figure displays untransformed values.",
            )
        )
    add_holm(tests, fig4d_indices)
    for i in fig4d_indices:
        tests[i]["notes"] += " Holm family: the four time-point comparisons in Fig. 4d."
    toxin = pd.read_excel(f4, sheet_name="Fig.4e", header=2)
    toxin = toxin[toxin.Samples.notna()].copy()
    toxin["day"] = toxin.Samples.str.extract(r"-(?:91|54)-(\d+)-")[0].astype(int)
    toxin["cellular_toxin_fmol_cell"] = (
        toxin.Total
        * toxin["LC-MS/MS extract volume (mL)"]
        * toxin["Dilution factor"]
        * 1e6
        / (toxin["Cell density (cells/ml)"] * toxin["Culture volume collected (mL)"])
    )
    fig3e_indices = []
    for day, sub in toxin.groupby("day"):
        fig3e_indices.append(len(tests))
        tests.append(
            t_row(
                "Fig. 4",
                "e",
                f"WT vs edited at day {day}",
                "total cellular toxin (fmol cell−1)",
                sub.loc[sub.Strain == WT, "cellular_toxin_fmol_cell"],
                sub.loc[sub.Strain == EDITED, "cellular_toxin_fmol_cell"],
                "Two-tailed Welch t-test",
                False,
                f4.name,
                "Fig.4e",
                notes="Independent biological replicates; the test uses untransformed cellular-toxin values.",
            )
        )
    add_holm(tests, fig3e_indices)
    for i in fig3e_indices:
        tests[i]["notes"] += " Holm family: the eight time-point comparisons in Fig. 4e."

    f5 = source_dir / "Source_data_fig5.xlsx"
    sxta = pd.read_excel(f5, sheet_name="Fig.5b")
    fig4b_indices = []
    for day, sub in sxta.groupby("day"):
        fig4b_indices.append(len(tests))
        tests.append(
            t_row(
                "Fig. 5",
                "b",
                f"WT vs edited at day {day}",
                "sxtA expression (FPKM)",
                sub.loc[sub.group == WT, "fpkm"],
                sub.loc[sub.group == EDITED, "fpkm"],
                "Two-tailed Welch t-test",
                False,
                f5.name,
                "Fig.5b",
            )
        )
    add_holm(tests, fig4b_indices)

    pst = pd.read_excel(f5, sheet_name="Fig.5c")
    fig4c_indices = []
    for gene, sub in pst.groupby("pst_gene", sort=False):
        fig4c_indices.append(len(tests))
        tests.append(
            t_row(
                "Fig. 5",
                "c",
                f"WT vs edited for {gene}",
                "FPKM pooled across four sampled days",
                sub.loc[sub.group == WT, "fpkm"],
                sub.loc[sub.group == EDITED, "fpkm"],
                "Two-tailed Welch t-test",
                False,
                f5.name,
                "Fig.5c",
            )
        )
    add_holm(tests, fig4c_indices)

    gene_p: list[dict] = []

    fed1 = source_dir / "Source_data_Extended_Data_Fig1.xlsx"
    fed7 = source_dir / "Source_data_Extended_Data_Fig7.xlsx"
    ed1c = pd.read_excel(fed1, sheet_name="ED_Fig.1c", header=2)
    ed1c_bio = ed1c.groupby(["Number of cells", "Samples"], as_index=False)["Ct Value"].mean()
    ed1c_bio["copy_number_per_cell"] = (
        10 ** ((39.67 - ed1c_bio["Ct Value"]) / 3.59) * 80 / ed1c_bio["Number of cells"]
    )
    for cells, sub in ed1c_bio.groupby("Number of cells"):
        mean, sd, n = mean_sd(sub.copy_number_per_cell)
        summaries.append(
            {
                "figure": "Extended Data Fig. 1",
                "panel": "c",
                "group": f"{cells:g} cells",
                "outcome": "sxtA4 copy number per cell",
                "mean": mean,
                "sd": sd,
                "n": n,
                "source_file": fed1.name,
                "source_sheet": "ED_Fig.1c",
                "notes": "Technical Ct replicates averaged per biological sample; ×80 corrects fourfold dilution and 1/20 template fraction.",
            }
        )

    ed7 = pd.read_excel(fed7, sheet_name="ED_Fig.7")
    ed7_rows = ed7[
        ["module", "GO.ID", "term", "timepoint", "up", "down", "DEG", "FDR", "FDR_bin"]
    ].to_dict(orient="records")

    panel_map = [
        ("Fig. 1", "a–c,e,f", "Fig.1a-b_base_comp; Fig.1a-b_SNPs; Fig.1c_sequences; Fig.1e_metrics; Fig.1f_pair_prob", "plot_figure1.py", "Computational predictions or descriptive source data"),
        ("Fig. 2", "d,f", "Fig.2d; Fig.2f", "plot_figure2.py", "Welch t-tests"),
        ("Fig. 3", "b–e", "Fig.3b-g_raw; Fig.3b-e", "plot_figure3.py", "Descriptive; no inferential tests"),
        ("Fig. 3", "f,g", "Fig.3b-g_raw; Fig.3f; Fig.3g", "plot_figure3.py", "ANOVA/Tukey and Welch t-tests"),
        ("Fig. 3", "h–j", "Fig.3h-j", "plot_figure3.py", "Welch t-test; Welch ANOVA/Games-Howell; descriptive composition"),
        ("Fig. 3", "k,l", "Fig.3k; Fig.3l", "plot_figure3.py", "Descriptive"),
        ("Fig. 4", "b,c", "Fig.4b; Fig.4c", "plot_figure4.py", "Descriptive; no inferential comparisons"),
        ("Fig. 4", "d", "Fig.4d", "plot_figure4.py", "Log10-scale Welch tests with Holm correction"),
        ("Fig. 4", "e", "Fig.4e", "plot_figure4.py", "Welch tests with Holm correction"),
        ("Fig. 5", "a", "Fig.5a", "plot_figure5.py", "Descriptive"),
        ("Fig. 5", "b", "Fig.5b", "plot_figure5.py", "Welch tests with Holm correction"),
        ("Fig. 5", "c", "Fig.5c", "plot_figure5.py", "Welch tests with Holm correction"),
        ("Fig. 5", "d", "Fig.5d", "plot_figure5.py", "DESeq2 gene-level tests"),
        ("Fig. 5", "e", "Fig.5e", "plot_figure5.py", "Descriptive; no additional hypothesis test"),
        ("Extended Data Fig. 1", "b–f", "ED_Fig.1b; ED_Fig.1c; ED_Fig.1d; ED_Fig.1e; ED_Fig.1f", "plot_extended_data.py", "Descriptive or computational predictions"),
        ("Extended Data Fig. 2", "a,b", "ED_Fig.2a; ED_Fig.2b", "plot_extended_data.py", "Descriptive heat maps"),
        ("Extended Data Fig. 6", "a–e", "ED_Fig.6a–e", "plot_extended_data.py", "PCA/Pearson/descriptive"),
        ("Extended Data Fig. 7", "", "ED_Fig.7", "plot_extended_data.py", "GO enrichment FDR"),
    ]
    panel_rows = [
        {
            "figure": fig,
            "panel": panel,
            "source_file": {
                "Extended Data Fig. 1": "Source_data_Extended_Data_Fig1.xlsx",
                "Extended Data Fig. 2": "Source_data_Extended_Data_Fig2.xlsx",
                "Extended Data Fig. 6": "Source_data_Extended_Data_Fig6.xlsx",
                "Extended Data Fig. 7": "Source_data_Extended_Data_Fig7.xlsx",
            }.get(fig, f"Source_data_fig{fig.split('.')[1].strip()}.xlsx"),
            "source_sheet": sheet,
            "script": script,
            "statistics": stat,
        }
        for fig, panel, sheet, script, stat in panel_map
    ]

    return {
        "panel_map": panel_rows,
        "tests": tests,
        "summaries": summaries,
        "fig5e_gene_p": gene_p,
        "ed7_go_fdr": ed7_rows,
        "checks": checks,
    }


def write_excel_results(result: dict, path: Path) -> None:
    columns = [
        "figure",
        "panel",
        "comparison",
        "outcome",
        "n_group_1",
        "n_group_2",
        "test",
        "multiplicity_adjustment",
        "statistic",
        "df",
        "p_raw",
        "p_adjusted",
        "p_reported",
        "figure_label",
        "source",
        "notes",
    ]
    rows = []
    for test in result["tests"]:
        rows.append(
            {
                "figure": test["figure"],
                "panel": test["panel"],
                "comparison": test["comparison"],
                "outcome": test["outcome"],
                "n_group_1": test["group_1_n"],
                "n_group_2": test["group_2_n"],
                "test": test["test"],
                "multiplicity_adjustment": test["correction"],
                "statistic": test["statistic"],
                "df": test["df"],
                "p_raw": test["p_raw"],
                "p_adjusted": test["p_adjusted"],
                "p_reported": test["p_used_for_display"],
                "figure_label": test["significance"],
                "source": f"{test['source_file']} / {test['source_sheet']}",
                "notes": test["notes"],
            }
        )
    frame = pd.DataFrame(rows, columns=columns)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="statistical_tests", index=False)
        sheet = writer.sheets["statistical_tests"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.outlinePr.summaryBelow = True

        header_fill = PatternFill("solid", fgColor="E7E6E6")
        thin_grey = Side(style="thin", color="D9D9D9")
        for cell in sheet[1]:
            cell.font = Font(name="Times New Roman", size=10, bold=True, color="000000")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_grey)
        sheet.row_dimensions[1].height = 30

        widths = {
            "A": 13,
            "B": 8,
            "C": 28,
            "D": 31,
            "E": 11,
            "F": 11,
            "G": 28,
            "H": 31,
            "I": 13,
            "J": 15,
            "K": 14,
            "L": 14,
            "M": 14,
            "N": 12,
            "O": 34,
            "P": 55,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        p_columns = {11, 12, 13}
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Times New Roman", size=9, color="000000")
                cell.alignment = Alignment(
                    vertical="top",
                    horizontal="center" if cell.column in {1, 2, 5, 6, 9, 10, 11, 12, 13, 14} else "left",
                    wrap_text=True,
                )
                cell.border = Border(bottom=Side(style="hair", color="EDEDED"))
                if cell.column in p_columns and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000000E+00"
            sheet.row_dimensions[row[0].row].height = 30


def write_results(source_dir: Path, output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    result = calculate_statistics(source_dir)
    json_path = output_dir / "statistics_results.json"
    excel_path = output_dir / "statistics_results.xlsx"
    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, allow_nan=False), encoding="utf-8")
    write_excel_results(result, excel_path)
    for name in ("checks.csv", "ed7_go_fdr.csv", "fig4e_gene_p.csv", "fig5e_gene_p.csv", "panel_map.csv", "summaries.csv", "tests.csv"):
        path = output_dir / name
        if path.exists():
            path.unlink()
    return json_path, excel_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Recalculate figure statistics from the source-data workbooks.")
    parser.add_argument("--source", type=Path, required=True, help="Folder containing the source-data workbooks.")
    parser.add_argument("--output", type=Path, required=True, help="Folder for tabulated results.")
    args = parser.parse_args()
    for path in write_results(args.source, args.output):
        print(path)


if __name__ == "__main__":
    main()
